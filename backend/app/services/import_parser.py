from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


@dataclass(slots=True)
class ParsedBillboardRow:
    serial_no: int
    ad_type: str
    company_name: str
    permit_date: date | None
    size_text: str | None
    display_address: str
    legal_dong: str | None


def _normalize_ad_type(raw: str) -> str:
    value = raw.strip()
    if value == "옥상전광":
        return "ROOFTOP_LED"
    if value == "벽면전광":
        return "WALL_LED"
    return "UNKNOWN"


def parse_billboards_xlsx(path: str) -> tuple[list[ParsedBillboardRow], int]:
    wb = load_workbook(Path(path), data_only=True)
    ws = wb[wb.sheetnames[0]]

    parsed: list[ParsedBillboardRow] = []
    failed = 0

    for row in ws.iter_rows(min_row=3, max_col=7, values_only=True):
        serial_no, ad_type, company, permit_date, size_text, address, legal_dong = row
        if not serial_no or not ad_type or not company or not address:
            failed += 1
            continue

        normalized_permit_date: date | None = None
        if isinstance(permit_date, datetime):
            normalized_permit_date = permit_date.date()
        elif isinstance(permit_date, date):
            normalized_permit_date = permit_date

        parsed.append(
            ParsedBillboardRow(
                serial_no=int(serial_no),
                ad_type=_normalize_ad_type(str(ad_type)),
                company_name=str(company).strip(),
                permit_date=normalized_permit_date,
                size_text=str(size_text).strip() if size_text else None,
                display_address=str(address).strip(),
                legal_dong=str(legal_dong).strip() if legal_dong else None,
            )
        )

    return parsed, failed
